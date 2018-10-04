import os
import re
import sys
import openpyxl

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'IG_Connect.settings')

import django
django.setup()

from inventory.models import Item

# Type is equipment or book

def addItem(name, count, itemtype):
    for i in range(count):
        i = Item()
        i.itemType = itemtype
        i.itemName = name
        i.save()

if __name__ == '__main__':
    wb = openpyxl.load_workbook('IG_Hardware_Available.xlsx')
    sheet1 = wb['Sheet1']
    for i in range(1, sheet1.max_row + 1):
        name = sheet1.cell(row =i, column = 1).value
        count = sheet1.cell(row =i, column = 2).value
        print name + str(count)
        addItem(name, count, "equipment")

        
    





